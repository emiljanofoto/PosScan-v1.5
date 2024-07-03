import tkinter as tk
from tkinter import ttk, messagebox
import openpyxl
from datetime import datetime
import subprocess

def load_data():
    path = "/Users/emiliofoto/Documents/tkinter-excel-appv2/data.xlsx"
    workbook = openpyxl.load_workbook(path)
    sheet = workbook.active

    # Clear current treeview
    for row in treeview.get_children():
        treeview.delete(row)

    # Get headers
    headers = [cell.value for cell in sheet[1]]
    treeview["columns"] = headers
    for header in headers:
        treeview.heading(header, text=header, anchor="center")
        treeview.column(header, anchor="center", width=100, stretch=tk.YES)  # Editable column widths

    # Load all values into treeview
    for row in sheet.iter_rows(min_row=2, values_only=True):
        treeview.insert('', 'end', values=row)

def search_data(event=None):
    query = search_entry.get().lower()

    path = "/Users/emiliofoto/Documents/tkinter-excel-appv2/data.xlsx"
    workbook = openpyxl.load_workbook(path)
    sheet = workbook.active

    # Clear current treeview
    for row in treeview.get_children():
        treeview.delete(row)

    found_item = None

    # Search and display matching rows
    for row in sheet.iter_rows(min_row=2, values_only=True):
        if any(query in str(cell).lower() for cell in row):
            item_id = treeview.insert('', 'end', values=row)
            if query == str(row[5]).lower():  # Assuming the 5th column for deletion match
                found_item = item_id

    if found_item:
        treeview.selection_set(found_item)
        root.after(100, delete_record)

def delete_record():
    selected_item = treeview.selection()
    if selected_item:
        record_values = treeview.item(selected_item, 'values')

        path = "/Users/emiliofoto/Documents/tkinter-excel-appv2/data.xlsx"
        workbook = openpyxl.load_workbook(path)
        sheet = workbook.active

        for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, values_only=False):
            if str(row[6].value) == str(record_values[6]):  # Match on the 5th column
                sheet.delete_rows(row[0].row)
                workbook.save(path)
                break

        treeview.delete(selected_item)
        delete_button.config(state=tk.DISABLED)

        # Show green message box
        show_message("Skanimi u fshi", "green")
        search_entry.delete(0, tk.END)
        search_entry.focus()
        load_data()

def show_message(message, color):
    message_label.config(text=message, background=color)
    message_label.grid(row=0, column=1, padx=20, pady=10, sticky="nw")
    root.after(2000, lambda: message_label.grid_forget())

def on_treeview_select(event):
    selected_item = treeview.selection()
    if selected_item:
        delete_button.config(state=tk.NORMAL)
    else:
        delete_button.config(state=tk.DISABLED)

def update_clock():
    global current_time  # Use a global variable to store the current time
    current_time = datetime.now().strftime("%d/%m/%Y %H:%M:%S")
    time_label.config(text=current_time)
    time_label.after(1000, update_clock)

def clear_search_entry(event):
    if search_entry.get() == "KERKO":
        search_entry.delete(0, tk.END)

def open_main_form():
    subprocess.Popen(['python', 'main.py'])
    root.destroy()  # Close the secondary window

root = tk.Tk()
root.title("PosSCAN v1.5")
root.geometry("1285x715")

style = ttk.Style(root)
root.tk.call("source", "forest-light.tcl")
root.tk.call("source", "forest-dark.tcl")
style.theme_use("forest-dark")

# Main frame
frame = ttk.Frame(root)
frame.pack(fill="both", expand=True)

# Date and time label at the top
time_label = ttk.Label(frame, text="", font=("Helvetica", 12))
time_label.grid(row=0, column=0, padx=20, pady=10, sticky="nw")
update_clock()  # Start updating the clock

# Search data frame in the middle
top_frame = ttk.LabelFrame(frame, text="TABELA PER KERKIM", width=800, height=100)
top_frame.grid(row=1, column=0, padx=30, pady=10, sticky="e")

topentry_width = 30
custom_font2 = ("Helvetica", 24)

search_entry = ttk.Entry(top_frame, width=topentry_width, font=custom_font2)
search_entry.insert(0, "KERKO")
search_entry.grid(row=0, column=0, padx=5, pady=(0, 5))

search_entry.bind("<FocusIn>", clear_search_entry)
search_entry.bind("<Return>", search_data)  # Bind Enter key to search function

search_button = ttk.Button(top_frame, text="KERKO", command=search_data)
search_button.grid(row=0, column=1, padx=5, pady=(0, 5))

# New frame for Open Main Form button
open_main_frame = ttk.Frame(frame)
open_main_frame.grid(row=1, column=1, padx=60, pady=20, sticky="e")

# Create a ttk.Button for Open Main Form
open_main_button = ttk.Button(open_main_frame, text="SKANIME", command=open_main_form)
open_main_button.pack(side=tk.LEFT, padx=30)

# Style adjustments to match the Search button
open_main_button.configure(style="Custom.TButton")

# Define a custom style for the button to match the Search button
style.configure("Custom.TButton", foreground=style.lookup("TButton", "foreground"),
                background=style.lookup("TButton", "background"),
                font=("Helvetica", 20), padding=style.lookup("TButton", "padding"),
                relief=style.lookup("TButton", "relief"))

separator = ttk.Separator(frame, orient="horizontal")
separator.grid(row=2, column=0, columnspan=2, padx=30, pady=10, sticky="ew")

# Treeview frame
treeview_frame = ttk.Frame(frame)
treeview_frame.grid(row=3, column=0, columnspan=2, padx=10, pady=10, sticky="nsew")

treeview = ttk.Treeview(treeview_frame, show='headings')
treeview.grid(row=0, column=0, sticky='nsew')

vsb = ttk.Scrollbar(treeview_frame, orient="vertical", command=treeview.yview)
vsb.grid(row=0, column=1, sticky='ns')
treeview.configure(yscrollcommand=vsb.set)

hsb = ttk.Scrollbar(treeview_frame, orient="horizontal", command=treeview.xview)
hsb.grid(row=1, column=0, sticky='ew')
treeview.configure(xscrollcommand=hsb.set)

# Delete button
delete_button = ttk.Button(frame, text="FSHIJE SKANIMIN", command=delete_record, state=tk.DISABLED)
delete_button.grid(row=4, column=0, columnspan=2, padx=10, pady=10, sticky="ew")

treeview.bind("<<TreeviewSelect>>", on_treeview_select)  # Enable delete button on selection

# Message label for displaying messages
message_label = ttk.Label(frame, text="", font=("Helvetica", 12))

load_data()

# Make the treeview and other elements expand to fill the available space
frame.rowconfigure(3, weight=1)
frame.columnconfigure(0, weight=1)
treeview_frame.rowconfigure(0, weight=1)
treeview_frame.columnconfigure(0, weight=1)

footer_frame = ttk.Frame(root)
footer_frame.pack(side=tk.BOTTOM, fill=tk.X)

left_footer_label = ttk.Label(footer_frame, text="          Programmed by: Emiljano Foto", font=("Helvetica", 12), foreground="gray")
left_footer_label.pack(side=tk.LEFT, padx=10)

right_footer_label = ttk.Label(footer_frame, text="Forschner Albania          ", font=("Helvetica", 12), foreground="gray")
right_footer_label.pack(side=tk.RIGHT, padx=10)


root.mainloop()
