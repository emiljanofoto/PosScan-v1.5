import tkinter as tk
from tkinter import ttk, messagebox
import openpyxl
from datetime import datetime
from PIL import Image, ImageTk
import subprocess

MAX_ROWS_DISPLAYED = 11

def load_data():
    path = "/Users/emiliofoto/Documents/PosScan/data.xlsx"
    workbook = openpyxl.load_workbook(path)
    sheet = workbook.active

    all_values = list(sheet.values)[1:]
    last_20_values = all_values[-MAX_ROWS_DISPLAYED:]

    for row in treeview.get_children():
        treeview.delete(row)

    for value_tuple in reversed(last_20_values):
        treeview.insert('', 0, values=value_tuple[:-1])
    
    update_sum_of_sasi()

def update_sum_of_sasi():
    path = "/Users/emiliofoto/Documents/PosScan/data.xlsx"
    workbook = openpyxl.load_workbook(path)
    sheet = workbook.active

    sum_sasi = sum(cell.value for cell in sheet['B'] if isinstance(cell.value, (int, float)))
    sum_label.config(text=f"SASIA TOTALE:   {sum_sasi}")

def lookup_harness(pozicion):
    path = "/Users/emiliofoto/Documents/PosScan/LISTAPN.xlsx"
    workbook = openpyxl.load_workbook(path)
    sheet = workbook.active

    for row in sheet.iter_rows(min_row=2, values_only=True):
        pos_value, pn_value = row
        if pos_value == pozicion:
            return pn_value
    return None

def is_duplicate_etiketa(etiketa):
    path = "/Users/emiliofoto/Documents/PosScan/data.xlsx"
    workbook = openpyxl.load_workbook(path)
    sheet = workbook.active

    for row in sheet.iter_rows(min_row=2, values_only=True):
        if row[5] == etiketa:
            return True
    return False

def validate_etiketa(etiketa):
    if is_duplicate_etiketa(etiketa):
        error_message_etiketa.config(text="Kjo etikete eshte skanuar njehere! (DUPLIKAT)")
        return False

    parts = etiketa.split('/')
    if len(parts) != 3 or not all(parts):
        error_message_etiketa.config(text="Skanoni etiketen sakte: '----/-----/----'")
        return False
    else:
        error_message_etiketa.config(text="")

        po_entry.config(state='normal')
        po_entry.delete(0, 'end')
        po_entry.insert(0, parts[0])
        po_entry.config(state='readonly')

        pozicion_entry.config(state='normal')
        pozicion_entry.delete(0, 'end')
        pozicion_entry.insert(0, parts[1])
        pozicion_entry.config(state='readonly')

        sasi_entry.config(state='normal')
        sasi_entry.delete(0, 'end')
        sasi_entry.insert(0, parts[2])
        sasi_entry.config(state='readonly')

        harness_value = lookup_harness(parts[1])
        if harness_value:
            harness_entry.config(state='normal')
            harness_entry.delete(0, 'end')
            harness_entry.insert(0, harness_value)
            harness_entry.config(state='readonly')
        else:
            harness_entry.config(state='normal')
            harness_entry.delete(0, 'end')
            harness_entry.insert(0, "Not found")
            harness_entry.config(state='readonly')

        return True

def validate_adresa(adresa):
    if not (adresa.startswith("*") or adresa.startswith("Adresa")):
        error_message_adresa.config(text="Skano sakte kodin e raftit!")
        return False
    else:
        error_message_adresa.config(text="")
        return True

def insert_row():
    global current_time
    etiketa = etiketa_entry.get()
    adresa = adresa_entry.get()

    if not validate_etiketa(etiketa):
        return

    if not validate_adresa(adresa):
        return

    try:
        sasi = int(sasi_entry.get())
    except ValueError:
        messagebox.showerror("Skano etiketen sakte(sasia)!")
        return

    harness = harness_entry.get()
    po = po_entry.get()
    pozicion = pozicion_entry.get()
    data_ora = current_time

    path = "/Users/emiliofoto/Documents/PosScan/data.xlsx"
    workbook = openpyxl.load_workbook(path)
    sheet = workbook.active

    sheet.insert_rows(2)

    row_values = [pozicion, sasi, harness, po, adresa, etiketa, data_ora]
    for col_num, value in enumerate(row_values, start=1):
        sheet.cell(row=2, column=col_num, value=value)

    workbook.save(path)

    treeview.insert('', 0, values=row_values[:-1])

    if len(treeview.get_children()) > MAX_ROWS_DISPLAYED:
        treeview.delete(treeview.get_children()[-1])

    clear_entries_and_error_message()
    update_sum_of_sasi()

def clear_entries_and_error_message():
    pozicion_entry.config(state='normal')
    pozicion_entry.delete(0, 'end')
    pozicion_entry.insert(0, "Pozicion")
    pozicion_entry.config(state='readonly')

    sasi_entry.config(state='normal')
    sasi_entry.delete(0, 'end')
    sasi_entry.insert(0, "Sasi")
    sasi_entry.config(state='readonly')

    harness_entry.config(state='normal')
    harness_entry.delete(0, 'end')
    harness_entry.insert(0, "Harness")
    harness_entry.config(state='readonly')

    po_entry.config(state='normal')
    po_entry.delete(0, 'end')
    po_entry.insert(0, "PO")
    po_entry.config(state='readonly')

    adresa_entry.delete(0, 'end')
    adresa_entry.insert(0, "Adresa")
    etiketa_entry.delete(0, 'end')
    etiketa_entry.insert(0, "Etiketa")

    error_message_etiketa.config(text="")
    error_message_adresa.config(text="")

def clear_etiketa_entry(event):
    if etiketa_entry.get() == "Etiketa":
        etiketa_entry.delete(0, tk.END)
    error_message_etiketa.config(text="")

def clear_adresa_entry(event):
    adresa_entry.delete(0, tk.END)
    error_message_adresa.config(text="")

def set_adresa_entry(event):
    if not adresa_entry.get():
        adresa_entry.insert(0, "Adresa")
    validate_adresa(adresa_entry.get())

def reset_entries(event):
    if etiketa_entry.get() == "":
        etiketa_entry.insert(0, "Etiketa")
    if adresa_entry.get() == "":
        adresa_entry.insert(0, "Adresa")

def update_clock():
    global current_time
    current_time = datetime.now().strftime("%d/%m/%Y %H:%M:%S")
    time_label.config(text=current_time)
    time_label.after(1000, update_clock)

def open_secondary_form():
    subprocess.Popen(['python', 'secondary.py'])
    root.destroy()  

def open_log_page():
    subprocess.Popen(['python', 'logpage.py'])
    root.destroy()  

root = tk.Tk()
root.title("PosSCAN v1.5")
root.geometry("1285x750")

style = ttk.Style(root)
root.tk.call("source", "forest-light.tcl")
root.tk.call("source", "forest-dark.tcl")
style.theme_use("forest-dark")

frame = ttk.Frame(root)
frame.pack(fill="both", expand=True)

top_frame = ttk.LabelFrame(frame, text="SKANO ETIKETEN", width=800, height=100)
top_frame.grid(row=0, column=0, padx=30, pady=10)

topentry_width = 20
custom_font2 = ("Helvetica", 35)

etiketa_entry = ttk.Entry(top_frame, width=topentry_width, font=custom_font2)
etiketa_entry.insert(0, "Etiketa")
etiketa_entry.grid(row=0, column=0, padx=5, pady=(0, 5))

error_message_etiketa = ttk.Label(top_frame, text="", foreground="red", font=("Helvetica", 15))
error_message_etiketa.grid(row=1, column=0, padx=5, pady=(0, 5))

etiketa_entry.bind("<FocusIn>", clear_etiketa_entry)
etiketa_entry.bind("<FocusOut>", lambda e: validate_etiketa(etiketa_entry.get()))
etiketa_entry.bind("<Return>", lambda e: adresa_entry.focus())

adresa_frame = ttk.LabelFrame(frame, text="ADRESA E RAFTIT", width=800, height=100)
adresa_frame.grid(row=1, column=0, padx=30, pady=10)

adresa_entry = ttk.Entry(adresa_frame, width=topentry_width, font=custom_font2)
adresa_entry.insert(0, "Adresa")
adresa_entry.grid(row=0, column=0, padx=5, pady=5)

error_message_adresa = ttk.Label(adresa_frame, text="", foreground="red", font=("Helvetica", 15))
error_message_adresa.grid(row=1, column=0, padx=5, pady=5)

adresa_entry.bind("<FocusIn>", clear_adresa_entry)
adresa_entry.bind("<FocusOut>", set_adresa_entry)
adresa_entry.bind("<Return>", lambda e: [insert_row(), etiketa_entry.focus()])

button_frame = ttk.Frame(frame)
button_frame.grid(row=2, column=0, padx=30, pady=10)

style.configure('Green.TButton', font=('Helvetica', 30))

button = ttk.Button(button_frame, text="RUAJ TE DHENAT", command=insert_row, width=24, style='Green.TButton')
button.grid(row=0, column=0, padx=5, pady=10)

separator = ttk.Separator(frame)
separator.grid(row=3, column=0, padx=30, pady=10, sticky="ew")

widgets_frame = ttk.LabelFrame(frame, text="TE DHENA E SKANIMIT", width=800, height=100)
widgets_frame.grid(row=4, column=0, padx=30, pady=10)

entry_width = 35
custom_font = ("Helvetica", 14)

pozicion_entry = ttk.Entry(widgets_frame, width=entry_width, font=custom_font, state='readonly')
pozicion_entry.grid(row=0, column=0, padx=5, pady=(0, 5))

sasi_entry = ttk.Entry(widgets_frame, width=entry_width, font=custom_font, state='readonly')
sasi_entry.grid(row=0, column=1, padx=5, pady=(0, 5))

harness_entry = ttk.Entry(widgets_frame, width=entry_width, font=custom_font, state='readonly')
harness_entry.grid(row=0, column=2, padx=5, pady=(0, 5))

po_entry = ttk.Entry(widgets_frame, width=entry_width, font=custom_font, state='readonly')
po_entry.grid(row=0, column=3, padx=5, pady=(0, 5))

treeview_frame = ttk.Frame(frame)
treeview_frame.grid(row=5, column=0, padx=10, pady=10, sticky="nsew")

treeview = ttk.Treeview(treeview_frame, columns=('Pozicion', 'Sasi', 'Harness', 'PO', 'Adresa', 'Etiketa'), show='headings')
treeview.grid(row=0, column=0, sticky='nsew')

vsb = ttk.Scrollbar(treeview_frame, orient="vertical", command=treeview.yview)
vsb.grid(row=0, column=1, sticky='ns')
treeview.configure(yscrollcommand=vsb.set)

treeview.heading('Pozicion', text='Pozicion', anchor="center")
treeview.heading('Sasi', text='Sasi', anchor="center")
treeview.heading('Harness', text='Harness', anchor="center")
treeview.heading('PO', text='PO', anchor="center")
treeview.heading('Adresa', text='Adresa', anchor="center")
treeview.heading('Etiketa', text='Etiketa', anchor="center")

treeview.column('Pozicion', anchor="center")
treeview.column('Sasi', anchor="center")
treeview.column('Harness', anchor="center")
treeview.column('PO', anchor="center")
treeview.column('Adresa', anchor="center")
treeview.column('Etiketa', anchor="center")

data_frame = ttk.LabelFrame(frame, text="TE DHENA MBI FIJET", width=100, height=100)
data_frame.grid(row=1, column=0, rowspan=2, padx=50, pady=10, sticky="nw")

sum_label = ttk.Label(data_frame, text="Total Sasi: 0", font=("Helvetica", 20))
sum_label.grid(row=0, column=0, padx=10, pady=10)

load_data()

frame.rowconfigure(5, weight=1)
frame.columnconfigure(0, weight=1)

time_label = ttk.Label(frame, text="", font=("Helvetica", 12))
time_label.grid(row=0, column=0, padx=20, pady=10, sticky="nw")
update_clock()

logo_path = "/Users/emiliofoto/Documents/PosScan/logo.png"
logo_image = Image.open(logo_path)
logo_photo = ImageTk.PhotoImage(logo_image)

logo_label = ttk.Label(frame, image=logo_photo)
logo_label.image = logo_photo
logo_label.grid(row=0, padx=20, pady=10, sticky="ne")

log_button = ttk.Button(frame, text="<- DIL", command=open_log_page, width=15, style='BigSquareRed.TButton')
log_button.grid(row=0, column=0, padx=20, pady=50, sticky="nw")

footer_frame = ttk.Frame(root)
footer_frame.pack(side=tk.BOTTOM, fill=tk.X)

style.configure('BigSquareRed.TButton', font=('Helvetica', 18, 'bold'))
button_width = 5
button_height = 1

secondary_button = ttk.Button(frame, text="HAP SKANIMET", command=open_secondary_form, width=button_width, style='BigSquareRed.TButton')
secondary_button.grid(row=1, column=0, padx=70, pady=0, ipadx=80, ipady=40, sticky="ne")

footer_frame = ttk.Frame(root)
footer_frame.pack(side=tk.BOTTOM, fill=tk.X)

left_footer_label = ttk.Label(footer_frame, text="          Programmed by: Emiljano Foto", font=("Helvetica", 12), foreground="gray")
left_footer_label.pack(side=tk.LEFT, padx=10)

right_footer_label = ttk.Label(footer_frame, text="Forschner Albania          ", font=("Helvetica", 12), foreground="gray")
right_footer_label.pack(side=tk.RIGHT, padx=10)

etiketa_entry.focus()

root.mainloop()
