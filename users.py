import tkinter as tk
from tkinter import ttk
import openpyxl
import subprocess

# Load the data from the Excel file
def load_data():
    path = "/Users/emiliofoto/Documents/PosScan/logs.xlsx"
    workbook = openpyxl.load_workbook(path)
    sheet = workbook.active
    data = []

    for row in sheet.iter_rows(min_row=2, values_only=True):
        if row[0] is not None and row[1] is not None:
            data.append((str(row[0]), row[1]))

    return data, path, workbook, sheet

# Save the data to the Excel file
def save_data():
    for row in treeview.get_children():
        treeview.delete(row)
    for id_, name in data:
        treeview.insert('', 'end', values=(id_, name))

# Add a new user to the Excel file and treeview
def add_user():
    user_id = id_entry.get()
    name = name_entry.get()
    if not user_id or not name:
        show_error_message("Gabim: Ploteso te dy tabelat!")
        return
    for item in data:
        if item[0] == user_id:
            show_error_message("Gabim: Kjo ID ekziston!")
            return
    data.append((user_id, name))
    save_to_excel()
    save_data()
    id_entry.delete(0, tk.END)
    name_entry.delete(0, tk.END)
    show_success_message("ID u shtua me sukses!")

# Delete a selected user from the Excel file and treeview
def delete_user():
    selected_item = treeview.selection()
    if not selected_item:
        show_error_message("Gabim: Selekto nje ID!")
        return
    for item in selected_item:
        item_id = treeview.item(item, 'values')[0]
        data[:] = [d for d in data if d[0] != item_id]
    save_to_excel()
    save_data()
    show_success_message("ID u fshi me sukses!")

# Save the updated data to the Excel file
def save_to_excel():
    sheet.delete_rows(2, sheet.max_row)
    for i, (id_, name) in enumerate(data, start=2):
        sheet.cell(row=i, column=1).value = id_
        sheet.cell(row=i, column=2).value = name
    workbook.save(path)

# Function to show error message for a short duration
def show_error_message(message):
    error_label.config(text=message, fg="white", bg="red")
    error_label.grid(row=4, column=0, columnspan=2, padx=10, pady=10, sticky="nsew")
    root.after(3000, clear_error_message)

# Function to show success message for a short duration
def show_success_message(message):
    error_label.config(text=message, fg="white", bg="green")
    error_label.grid(row=4, column=0, columnspan=2, padx=10, pady=10, sticky="nsew")
    root.after(3000, clear_error_message)

# Function to clear the error message label
def clear_error_message():
    error_label.grid_forget()

# Function to open the log page form and close the current form
def open_logpage_form():
    root.destroy()  # Close the current form
    subprocess.Popen(['python', 'logpage.py'])

# Main application window
root = tk.Tk()
root.title("Shtimi i Perdorusave")
root.geometry("420x500")

# Load data from Excel
data, path, workbook, sheet = load_data()

# ID entry
ttk.Label(root, text="ID:").grid(row=0, column=0, padx=10, pady=10, sticky=tk.W)
id_entry = ttk.Entry(root, font=("Helvetica", 14))
id_entry.grid(row=0, column=1, padx=10, pady=10)

# Name entry
ttk.Label(root, text="Emer:").grid(row=1, column=0, padx=10, pady=10, sticky=tk.W)
name_entry = ttk.Entry(root, font=("Helvetica", 14))
name_entry.grid(row=1, column=1, padx=10, pady=10)

# Add button
add_button = ttk.Button(root, text="Shto", command=add_user)
add_button.grid(row=2, column=0, padx=10, pady=10)

# Delete button
delete_button = ttk.Button(root, text="Fshij", command=delete_user)
delete_button.grid(row=2, column=1, padx=10, pady=10)

# Treeview to display users
treeview_frame = ttk.Frame(root)
treeview_frame.grid(row=3, column=0, columnspan=2, padx=10, pady=10, sticky='nsew')

treeview = ttk.Treeview(treeview_frame, columns=("ID", "Name"), show="headings")
treeview.heading("ID", text="ID")
treeview.heading("Name", text="Name")
treeview.pack(fill="both", expand=True)

vsb = ttk.Scrollbar(treeview_frame, orient="vertical", command=treeview.yview)
vsb.pack(side=tk.RIGHT, fill=tk.Y)
treeview.configure(yscrollcommand=vsb.set)

hsb = ttk.Scrollbar(treeview_frame, orient="horizontal", command=treeview.xview)
hsb.pack(side=tk.BOTTOM, fill=tk.X)
treeview.configure(xscrollcommand=hsb.set)

# Error message label
error_label = tk.Label(root, text="", fg="white", bg=root.cget("bg"), font=("Helvetica", 12))
error_label.grid(row=4, column=0, columnspan=2, padx=10, pady=10, sticky="nsew")

# Close and open log page button
close_button = ttk.Button(root, text="        MBYLL FAQEN         ", command=open_logpage_form)
close_button.grid(row=7, column=0, padx=0, pady=10, sticky=tk.NE)

# Load the data into the treeview
save_data()

root.mainloop()
